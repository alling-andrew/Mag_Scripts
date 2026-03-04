#!/usr/bin/env python3
"""
psend_logger.py

Reads lines like:
  $psend,<time>,<sensor type>,<sensor index>,<value1>,...<valueN>*<checksum>

- Validates checksum (XOR of chars between '$' and '*', compared to hex after '*')
- Logs parsed data to CSV in ~/Documents (Windows)
- On Ctrl+C, optionally converts CSV -> XLSX

Install:
  pip install pyserial openpyxl
"""

from __future__ import annotations

import csv
import datetime as dt
import time
from pathlib import Path

import serial

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # XLSX conversion optional


# -------------------- Config --------------------
PORT = "/dev/ttyUSB0"
BAUDRATE = 115200

BYTESIZE = serial.EIGHTBITS
PARITY = serial.PARITY_NONE
STOPBITS = serial.STOPBITS_ONE

XONXOFF = False
RTSCTS = False
DSRDTR = False

READ_TIMEOUT_S = 0.1  # serial read timeout

### raw logging ###

LOG_RAW = False

FLUSH_EVERY = 500


# -------------------- Helpers --------------------
def nmea_xor_checksum(payload: str) -> int:
    """XOR checksum over the payload string (no $ or *)."""
    c = 0
    for ch in payload:
        c ^= ord(ch)
    return c


def parse_psend_line(line: str) -> dict | None:
    """
    Returns a dict with parsed fields or None if it doesn't look like psend.
    Dict keys:
      recv_iso, raw, checksum_ok, msg_time, sensor_type, sensor_index, values(list[str])
    """
    s = line.strip()
    if not s.startswith("$PSEND"):
        return None

###logging time with UTC###
    recv_ts = time.time()

    # Split checksum
    if "*" not in s:
        return {
            "recv_ts": recv_ts,
            "raw": s,
            "checksum_ok": False,
            "msg_time": "",
            "sensor_type": "",
            "sensor_index": "",
            "values": [],
        }

    body, chk_str = s.split("*", 1)
    chk_str = chk_str.strip()

    # body includes the leading '$'
    payload = body[1:]  # remove '$'

    # Validate checksum: typically 2 hex chars, but handle longer/shorter gracefully
    checksum_ok = False
    given_chk = None
    try:
        # take first 2 hex characters if present
        given_chk = int(chk_str[:2], 16)
        calc_chk = nmea_xor_checksum(payload)
        checksum_ok = (given_chk == calc_chk)
    except Exception:
        checksum_ok = False

    # Parse CSV-like fields from payload
    # payload: "psend,<time>,<sensor type>,<sensor index>,<value1>,..."
    parts = payload.split(",")
    # parts[0] should be "psend"
    msg_time = parts[1].strip() if len(parts) > 1 else ""
    sensor_type = parts[2].strip() if len(parts) > 2 else ""
    sensor_index = parts[3].strip() if len(parts) > 3 else ""
    values = [p.strip() for p in parts[4:]] if len(parts) > 4 else []

    return {
        ### unix ts ###
        "recv_ts": recv_ts,
        "raw": s,
        "checksum_ok": checksum_ok,
        "msg_time": msg_time,
        "sensor_type": sensor_type,
        "sensor_index": sensor_index,
        "values": values,
    }


def csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("data")

    with csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(xlsx_path)


# -------------------- Main --------------------
def main():
    documents_dir = Path.home() / "Documents/Mag_Scripts/Mag_Data"
    documents_dir.mkdir(parents=True, exist_ok=True)

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = documents_dir / f"psend_log_{stamp}.csv"
    xlsx_path = documents_dir / f"psend_log_{stamp}.xlsx"

    # CSV columns:
    ###recv_iso, checksum_ok, msg_time, sensor_type, sensor_index, values_joined, raw###
    header = [
        ###unix time###
        "timestamp_unix",
        "checksum_ok",
        "msg_time",
        "sensor_type",
        "sensor_index",
        "values",   # semicolon-separated to support variable count
        "raw",
    ]

    print(f"[INFO] Logging to: {csv_path}")
    print("[INFO] Press Ctrl+C to stop.\n")

    # Open serial
    ser = serial.Serial(
        port=PORT,
        baudrate=BAUDRATE,
        bytesize=BYTESIZE,
        parity=PARITY,
        stopbits=STOPBITS,
        timeout=READ_TIMEOUT_S,
        xonxoff=XONXOFF,
        rtscts=RTSCTS,
        dsrdtr=DSRDTR,
    )

    lines_ok = 0
    lines_bad = 0
    lines_other = 0

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        f.flush()

        try:
            while True:
                raw_bytes = ser.readline()
                if not raw_bytes:
                    continue

                try:
                    line = raw_bytes.decode("utf-8", errors="replace").strip()
                except Exception:
                    line = str(raw_bytes)

                parsed = parse_psend_line(line)
                if parsed is None:
                    lines_other += 1
                    continue

                if parsed["checksum_ok"]:
                    lines_ok += 1
                else:
                    lines_bad += 1

                writer.writerow([
                    parsed["recv_ts"],
                    "TRUE" if parsed["checksum_ok"] else "FALSE",
                    parsed["msg_time"],
                    parsed["sensor_type"],
                    parsed["sensor_index"],
                    ";".join(parsed["values"]),
                    parsed["raw"] if LOG_RAW else "",
                ])

                # periodic flush so you don't lose data if stopped
                if (lines_ok + lines_bad) % FLUSH_EVERY == 0:
                    f.flush()
                    print(f"[INFO] psend logged: ok={lines_ok} bad_chk={lines_bad} other={lines_other}")

        except KeyboardInterrupt:
            print("\n[INFO] Stopping...")
        finally:
            try:
                ser.close()
            except Exception:
                pass

    print(f"[INFO] Final counts: ok={lines_ok}, bad_chk={lines_bad}, other={lines_other}")
    print(f"[INFO] CSV saved: {csv_path}")

    # Optional XLSX conversion
    try:
        if Workbook is None:
            print("[INFO] openpyxl not installed, skipping XLSX conversion.")
            print("       Install it with: pip install openpyxl")
        else:
            csv_to_xlsx(csv_path, xlsx_path)
            print(f"[INFO] XLSX saved: {xlsx_path}")
    except Exception as e:
        print(f"[WARN] XLSX conversion failed: {e}")


if __name__ == "__main__":
    main()

