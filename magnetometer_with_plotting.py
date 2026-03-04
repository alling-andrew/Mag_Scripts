#!/usr/bin/env python3
"""
psend_logger.py

Reads lines like:
  $PSEND,<time>,<sensor type>,<sensor index>,<value1>,...<valueN>*<checksum>

- Validates checksum (XOR of chars between '$' and '*', compared to hex after '*')
- Logs parsed data to CSV in ~/Documents (Windows)
- Live plots:
    1) x/y/z components when sensor_type == '4'
    2) total magnetic intensity when sensor_type != '4' (scalar)
- On Ctrl+C, optionally converts CSV -> XLSX

Install:
  pip install pyserial matplotlib openpyxl
"""

from __future__ import annotations

import csv
import datetime as dt
import time
from collections import deque
from pathlib import Path

import serial
import matplotlib.pyplot as plt

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # XLSX conversion optional


# -------------------- Config --------------------
PORT = "/dev/ttyUSB0"
BAUDRATE = 115200

BYTESIZE = serial.EIGHTBITS
PARITY = serial.PARITY_NONE
STOPBITS = serial.STOPBITS_ONE

XONXOFF = False
RTSCTS = False
DSRDTR = False

READ_TIMEOUT_S = 0.1  # serial read timeout

# Live plot tuning
PLOT_HZ = 10.0               # refresh rate (Hz)
PLOT_PERIOD = 1.0 / PLOT_HZ
MAX_POINTS = 350            # rolling history

LOG_RAW = False

FLUSH_EVERY = 500

# -------------------- Helpers --------------------
def nmea_xor_checksum(payload: str) -> int:
    """XOR checksum over the payload string (no $ or *)."""
    c = 0
    for ch in payload:
        c ^= ord(ch)
    return c


def _safe_float(s):
    try:
        return float(s)
    except Exception:
        return None


def parse_psend_line(line: str) -> dict | None:
    """
    Returns a dict with parsed fields or None if it doesn't look like psend.
    Dict keys:
      recv_iso, raw, checksum_ok, msg_time, sensor_type, sensor_index, values(list[str])
    """
    s = line.strip()
    if not s.startswith("$PSEND"):
        return None
###logging time with UTC###:
    recv_iso = dt.datetime.now().isoformat(timespec="milliseconds")

    # Split checksum
    if "*" not in s:
        return {
            "recv_iso": recv_iso,
            "raw": s,
            "checksum_ok": False,
            "msg_time": "",
            "sensor_type": "",
            "sensor_index": "",
            "values": [],
            "value_x": None,
            "value_y": None,
            "value_z": None,
        }

    body, chk_str = s.split("*", 1)
    chk_str = chk_str.strip()

    # body includes the leading '$'
    payload = body[1:]  # remove '$'

    # Validate checksum
    checksum_ok = False
    try:
        given_chk = int(chk_str[:2], 16)
        calc_chk = nmea_xor_checksum(payload)
        checksum_ok = (given_chk == calc_chk)
    except Exception:
        checksum_ok = False

    # Parse CSV-like fields from payload
    parts = payload.split(",")
    msg_time = parts[1].strip() if len(parts) > 1 else ""
    sensor_type = parts[2].strip() if len(parts) > 2 else ""
    sensor_index = parts[3].strip() if len(parts) > 3 else ""

    if sensor_type == '4':
        values = ''
        dir_values = [p.split(';') for p in parts[4:]] if len(parts) > 4 else []
        # your expected structure: [ [x], [y], [z] ]
        value_x = dir_values[0][0].strip() if len(dir_values) > 0 and len(dir_values[0]) > 0 else None
        value_y = dir_values[1][0].strip() if len(dir_values) > 1 and len(dir_values[1]) > 0 else None
        value_z = dir_values[2][0].strip() if len(dir_values) > 2 and len(dir_values[2]) > 0 else None
    else:
        values = [p.strip() for p in parts[4:]] if len(parts) > 4 else []
        value_x = None
        value_y = None
        value_z = None

    return {
        "recv_iso": recv_iso,
        "raw": s,
        "checksum_ok": checksum_ok,
        "msg_time": msg_time,
        "sensor_type": sensor_type,
        "sensor_index": sensor_index,
        "value_total": values,
        'value_x': value_x,
        'value_y': value_y,
        'value_z': value_z,
    }


def csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("data")

    with csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(xlsx_path)


# -------------------- Plot setup --------------------
def setup_live_plots():
    plt.ion()

    # Vector plot
    fig_vec, ax_vec = plt.subplots()
    ax_vec.set_title("Magnetic Vector (X, Y, Z)  [sensor_type = 4]")
    ax_vec.set_xlabel("sample #")
    ax_vec.set_ylabel("value")
    line_x, = ax_vec.plot([], [], label="X")
    line_y, = ax_vec.plot([], [], label="Y")
    line_z, = ax_vec.plot([], [], label="Z")
    ax_vec.legend(loc="upper right")

    # Scalar plot
    fig_tot, ax_tot = plt.subplots()
    ax_tot.set_title("Total Magnetic Intensity  [sensor_type = 3]")
    ax_tot.set_xlabel("sample #")
    ax_tot.set_ylabel("total")
    line_tot, = ax_tot.plot([], [], label="Total")
    ax_tot.legend(loc="upper right")

    fig_vec.show()
    fig_tot.show()

    return (fig_vec, ax_vec, line_x, line_y, line_z), (fig_tot, ax_tot, line_tot)


def _autoscale(ax, y_values):
    if not y_values:
        return
    ymin = min(y_values)
    ymax = max(y_values)
    if ymin == ymax:
        ymin -= 1.0
        ymax += 1.0
    ax.set_ylim(ymin, ymax)


# -------------------- Main --------------------
def main():
    documents_dir = Path.home() / "Documents/Mag_Scripts/Mag_Data"
    documents_dir.mkdir(parents=True, exist_ok=True)

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = documents_dir / f"psend_log_{stamp}.csv"
    xlsx_path = documents_dir / f"psend_log_{stamp}.xlsx"

    header = [
        "recv_iso",
        "checksum_ok",
        "msg_time",
        "sensor_type",
        "sensor_index",
        "value total",
        'value_x',
        'value_y',
        'value_z',
        "raw",
    ]

    print(f"[INFO] Logging to: {csv_path}")
    print("[INFO] Live plotting enabled (throttled). Press Ctrl+C to stop.\n")

    # Rolling buffers for plots
    idx_vec = deque(maxlen=MAX_POINTS)
    x_buf = deque(maxlen=MAX_POINTS)
    y_buf = deque(maxlen=MAX_POINTS)
    z_buf = deque(maxlen=MAX_POINTS)

    idx_tot = deque(maxlen=MAX_POINTS)
    tot_buf = deque(maxlen=MAX_POINTS)

    vec_plot, tot_plot = setup_live_plots()
    fig_vec, ax_vec, line_x, line_y, line_z = vec_plot
    fig_tot, ax_tot, line_tot = tot_plot

    # Open serial
    ser = serial.Serial(
        port=PORT,
        baudrate=BAUDRATE,
        bytesize=BYTESIZE,
        parity=PARITY,
        stopbits=STOPBITS,
        timeout=READ_TIMEOUT_S,
        xonxoff=XONXOFF,
        rtscts=RTSCTS,
        dsrdtr=DSRDTR,
    )

    lines_ok = 0
    lines_bad = 0
    lines_other = 0

    sample_vec = 0
    sample_tot = 0

    last_plot = time.perf_counter()

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        f.flush()

        try:
            while True:
                raw_bytes = ser.readline()

                # Keep GUI responsive even if no serial data
                if not raw_bytes:
                    plt.pause(0.001)
                    continue

                try:
                    line = raw_bytes.decode("utf-8", errors="replace").strip()
                except Exception:
                    line = str(raw_bytes)

                parsed = parse_psend_line(line)
                if parsed is None:
                    lines_other += 1
                    continue

                if parsed["checksum_ok"]:
                    lines_ok += 1
                    writer.writerow([
                    parsed["recv_iso"],
                    "TRUE" if parsed["checksum_ok"] else "FALSE",
                    parsed["msg_time"],
                    parsed["sensor_type"],
                    parsed["sensor_index"],
                    ";".join(parsed["value_total"]) if isinstance(parsed["value_total"], list) else parsed["value_total"],
                    parsed['value_x'],
                    parsed['value_y'],
                    parsed['value_z'],
                    parsed["raw"],
                ])
                else:
                    lines_bad += 1



                # -------- Buffer updates for live plots --------
                if parsed["sensor_type"] == "4":
                    vx = _safe_float(parsed["value_x"])
                    vy = _safe_float(parsed["value_y"])
                    vz = _safe_float(parsed["value_z"])
                    if vx is not None and vy is not None and vz is not None:
                        sample_vec += 1
                        idx_vec.append(sample_vec)
                        x_buf.append(vx)
                        y_buf.append(vy)
                        z_buf.append(vz)
                elif parsed["sensor_type"] == "3":
                    # total intensity expected in value_total list
                    vt = None
                    if isinstance(parsed["value_total"], list) and len(parsed["value_total"]) > 0:
                        vt = _safe_float(parsed["value_total"][0])
                    if vt is not None:
                        sample_tot += 1
                        idx_tot.append(sample_tot)
                        tot_buf.append(vt)

                # -------- Throttled plot refresh --------
                now = time.perf_counter()
                if now - last_plot >= PLOT_PERIOD:
                    last_plot = now

                    # Vector plot
                    if len(idx_vec) > 1:
                        xv = list(idx_vec)
                        line_x.set_data(xv, list(x_buf))
                        line_y.set_data(xv, list(y_buf))
                        line_z.set_data(xv, list(z_buf))

                        ax_vec.set_xlim(xv[0], xv[-1])
                        _autoscale(ax_vec, list(x_buf) + list(y_buf) + list(z_buf))

                        fig_vec.canvas.draw_idle()
                        fig_vec.canvas.flush_events()

                    # Total plot
                    if len(idx_tot) > 1:
                        xt = list(idx_tot)
                        line_tot.set_data(xt, list(tot_buf))

                        ax_tot.set_xlim(xt[0], xt[-1])
                        _autoscale(ax_tot, list(tot_buf))

                        fig_tot.canvas.draw_idle()
                        fig_tot.canvas.flush_events()

                    plt.pause(0.001)

                # periodic flush so you don't lose data if stopped
                if (lines_ok + lines_bad) % 200 == 0:
                    f.flush()
                    print(f"[INFO] psend logged: ok={lines_ok} bad_chk={lines_bad} other={lines_other}")

        except KeyboardInterrupt:
            print("\n[INFO] Stopping...")
            
        finally:
            try:
                ser.close()
            except Exception:
                pass

    print(f"[INFO] Final counts: ok={lines_ok}, bad_chk={lines_bad}, other={lines_other}")
    print(f"[INFO] CSV saved: {csv_path}")

    # Optional XLSX conversion
    try:
        if Workbook is None:
            print("[INFO] openpyxl not installed, skipping XLSX conversion.")
            print("       Install it with: pip install openpyxl")
        else:
            csv_to_xlsx(csv_path, xlsx_path)
            print(f"[INFO] XLSX saved: {xlsx_path}")
    except Exception as e:
        print(f"[WARN] XLSX conversion failed: {e}")


if __name__ == "__main__":
    main()
